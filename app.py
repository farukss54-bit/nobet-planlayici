import streamlit as st
import pandas as pd
from datetime import datetime
from ortools.sat.python import cp_model
import io
import csv
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
import holidays


def calc_month_days(yil, ay):
    """Ayın gün sayısını verir"""
    if ay == 12:
        return (datetime(yil + 1, 1, 1) - datetime(yil, 12, 1)).days
    return (datetime(yil, ay + 1, 1) - datetime(yil, ay, 1)).days


def parse_day_numbers(text: str, max_day: int):
    """Parse day numbers like: 1,5,12 or 1-5,12"""
    if not text or not text.strip():
        return set()
    out = set()
    parts = [p.strip() for p in text.split(',') if p.strip()]
    for p in parts:
        if '-' in p:
            try:
                a, b = p.split('-', 1)
                a = int(a.strip())
                b = int(b.strip())
                if a > b:
                    a, b = b, a
                for d in range(a, b + 1):
                    if 1 <= d <= max_day:
                        out.add(d)
            except:
                pass
        else:
            try:
                d = int(p)
                if 1 <= d <= max_day:
                    out.add(d)
            except:
                pass
    return out


def get_turkish_holidays(year, month):
    """Türkiye'deki resmi tatilleri otomatik çeker"""
    try:
        tr_holidays = holidays.Turkey(years=year)
        return {date.day: name for date, name in tr_holidays.items() if date.month == month}
    except:
        return {}


def solve_schedule(yil, ay, personeller, target, izinler, holidays, no_pairs, want_pairs, prefer_map=None, soft_no_pairs=None, w_gap3=300, gap2_weight=1000):
    """CP-SAT Solver"""
    if prefer_map is None:
        prefer_map = {}
    if soft_no_pairs is None:
        soft_no_pairs = []

    # Ağırlıklar / kurallar
    MIN_PEOPLE_PER_DAY = 1
    MAX_GUNASIRI_PER_PERSON = 1

    W_DAILY_EQUAL = 5000
    W_FRI = 1000
    W_SAT = 1000
    W_SUN = 1000
    W_HOL = 200
    W_WEEKENDLIKE = 50
    W_PAIR_REWARD = 30
    W_GAP3 = w_gap3
    W_SOFT_NO_PAIR = 800
    PREF_WEIGHT = 2

    gun_sayisi = calc_month_days(yil, ay)
    nP = len(personeller)
    if nP == 0:
        raise ValueError("Personel listesi boş.")

    # Basit feasibility
    total_target = sum(int(target[p]) for p in personeller)
    if total_target < gun_sayisi:
        raise ValueError(f"İmkânsız: toplam hedef ({total_target}) < gün sayısı ({gun_sayisi}).")

    def weekday(d):
        return datetime(yil, ay, d).weekday()

    def days_by_weekday(wd):
        return [d for d in range(1, gun_sayisi + 1) if weekday(d) == wd]

    # Model
    model = cp_model.CpModel()
    x = {}
    for p in range(nP):
        for d in range(1, gun_sayisi + 1):
            x[p, d] = model.NewBoolVar(f"x_{p}_{d}")

    # Günlük toplam kişi
    day_total = {}
    for d in range(1, gun_sayisi + 1):
        s = model.NewIntVar(0, nP, f"day_total_{d}")
        model.Add(s == sum(x[p, d] for p in range(nP)))
        model.Add(s >= MIN_PEOPLE_PER_DAY)
        day_total[d] = s

    # Hedef nöbet sayıları (hard exact)
    for p, name in enumerate(personeller):
        model.Add(sum(x[p, d] for d in range(1, gun_sayisi + 1)) == int(target[name]))

    # İzinler (hard)
    for p, name in enumerate(personeller):
        for d in izinler.get(name, set()):
            if 1 <= d <= gun_sayisi:
                model.Add(x[p, d] == 0)

    # Ardışık gün yok (hard)
    for p in range(nP):
        for d in range(1, gun_sayisi):
            model.Add(x[p, d] + x[p, d + 1] <= 1)

    # Günaşırı (gap=2) kişi başı max 1 (hard)
    for p in range(nP):
        gap2 = []
        for d in range(1, gun_sayisi - 1):
            b = model.NewBoolVar(f"gap2_{p}_{d}")
            model.Add(b <= x[p, d])
            model.Add(b <= x[p, d + 2])
            model.Add(b >= x[p, d] + x[p, d + 2] - 1)
            gap2.append(b)
        model.Add(sum(gap2) <= MAX_GUNASIRI_PER_PERSON)

    # Birlikte tutamasın (hard)
    name_to_idx = {n: i for i, n in enumerate(personeller)}
    for (a, b) in no_pairs:
        if a not in name_to_idx or b not in name_to_idx:
            continue
        pa = name_to_idx[a]
        pb = name_to_idx[b]
        for d in range(1, gun_sayisi + 1):
            model.Add(x[pa, d] + x[pb, d] <= 1)

    objective_terms = []

    # SOFT #1: Günlük kişi sayısı eşitliği
    totals = [day_total[d] for d in range(1, gun_sayisi + 1)]
    mn_day = model.NewIntVar(0, nP, "mn_day")
    mx_day = model.NewIntVar(0, nP, "mx_day")
    model.AddMinEquality(mn_day, totals)
    model.AddMaxEquality(mx_day, totals)
    diff_day = model.NewIntVar(0, nP, "diff_day")
    model.Add(diff_day == mx_day - mn_day)
    objective_terms.append(diff_day * W_DAILY_EQUAL)

    # Cuma/Cts/Paz + tatil adaleti
    fri_days = days_by_weekday(4)
    sat_days = days_by_weekday(5)
    sun_days = days_by_weekday(6)
    holiday_days = sorted(list(holidays)) if holidays else []
    weekend_like_days = sorted(list(set(fri_days + sat_days + sun_days) | set(holiday_days)))

    def add_fairness(days_set, weight, tag):
        if not days_set:
            return
        counts = []
        for p in range(nP):
            c = model.NewIntVar(0, len(days_set), f"{tag}_cnt_{p}")
            model.Add(c == sum(x[p, d] for d in days_set))
            counts.append(c)
        mn = model.NewIntVar(0, len(days_set), f"{tag}_min")
        mx = model.NewIntVar(0, len(days_set), f"{tag}_max")
        model.AddMinEquality(mn, counts)
        model.AddMaxEquality(mx, counts)
        diff = model.NewIntVar(0, len(days_set), f"{tag}_diff")
        model.Add(diff == mx - mn)
        objective_terms.append(diff * weight)

    add_fairness(fri_days, W_FRI, "fri")
    add_fairness(sat_days, W_SAT, "sat")
    add_fairness(sun_days, W_SUN, "sun")
    add_fairness(holiday_days, W_HOL, "hol")
    add_fairness(weekend_like_days, W_WEEKENDLIKE, "wklike")

    # SOFT #2: Gap=3 (2 Gün Boşluk Tercihi)
    for p in range(nP):
        for d in range(1, gun_sayisi - 1):
            g3 = model.NewBoolVar(f"soft_gap3_{p}_{d}")
            model.Add(g3 >= x[p, d] + x[p, d + 2] - 1)
            objective_terms.append(g3 * W_GAP3)

    # Birlikte tutsun: hard min + soft ödül
    for (a, b, min_k) in want_pairs:
        if a not in name_to_idx or b not in name_to_idx:
            continue
        pa = name_to_idx[a]
        pb = name_to_idx[b]

        together_bools = []
        for d in range(1, gun_sayisi + 1):
            t = model.NewBoolVar(f"together_{pa}_{pb}_{d}")
            model.Add(t <= x[pa, d])
            model.Add(t <= x[pb, d])
            model.Add(t >= x[pa, d] + x[pb, d] - 1)
            together_bools.append(t)

        together_cnt = model.NewIntVar(0, gun_sayisi, f"together_cnt_{pa}_{pb}")
        model.Add(together_cnt == sum(together_bools))
        model.Add(together_cnt >= int(min_k))
        objective_terms.append(together_cnt * (-W_PAIR_REWARD))

    # SOFT #3: Esnek Uyuşmazlık (Soft No-Pair)
    for (a, b) in soft_no_pairs:
        if a in name_to_idx and b in name_to_idx:
            pa, pb = name_to_idx[a], name_to_idx[b]
            for d in range(1, gun_sayisi + 1):
                together = model.NewBoolVar(f"soft_together_{pa}_{pb}_{d}")
                model.Add(together >= x[pa, d] + x[pb, d] - 1)
                objective_terms.append(together * W_SOFT_NO_PAIR)

    # Prefer edilen günler: denk geldikçe ödül
    pref_hits = []
    for p_idx, p_name in enumerate(personeller):
        for d in prefer_map.get(p_name, set()):
            if 1 <= d <= gun_sayisi:
                pref_hits.append(x[p_idx, d])

    if pref_hits:
        objective_terms.append(-PREF_WEIGHT * sum(pref_hits))

    model.Minimize(sum(objective_terms))

    # Solve
    solver = cp_model.CpSolver()
    solver.parameters.max_time_in_seconds = 60.0
    solver.parameters.num_search_workers = 8

    status = solver.Solve(model)
    if status not in (cp_model.OPTIMAL, cp_model.FEASIBLE):
        raise ValueError("Çözüm bulunamadı (kısıtlar fazla sıkı olabilir).")

    schedule = {}
    for d in range(1, gun_sayisi + 1):
        schedule[d] = []
        for p, name in enumerate(personeller):
            if solver.Value(x[p, d]) == 1:
                schedule[d].append(name)

    return schedule


def diagnose_no_solution(yil, ay, personeller, target_map, izinler, holidays, want_pairs, no_pairs, min_staff_per_day=1, max_staff_per_day=3):
    """CP-SAT 'neden' söylemez. Biz hızlı mantık kontrolleri yapıp olası nedenleri üretiriz."""
    problems = []
    gun_sayisi = calc_month_days(yil, ay)
    holidays_set = set(holidays or [])

    # 1) Gün bazında müsaitlik
    avail_by_day = {}
    for d in range(1, gun_sayisi + 1):
        avail = []
        for p in personeller:
            if d not in izinler.get(p, set()):
                avail.append(p)
        avail_by_day[d] = avail

    # Günlerde minimum personel sağlanıyor mu?
    zero_days = [d for d in range(1, gun_sayisi + 1) if len(avail_by_day[d]) < min_staff_per_day]
    if zero_days:
        problems.append(
            f"En az {min_staff_per_day} kişi şartını sağlayamayan günler var: {zero_days[:10]}"
            + (" ..." if len(zero_days) > 10 else "")
        )

    # 2) Toplam slot kapasitesi kontrolü
    total_required = sum(target_map.get(p, 0) for p in personeller)
    total_capacity = sum(min(len(avail_by_day[d]), max_staff_per_day) for d in range(1, gun_sayisi + 1))
    if total_required > total_capacity:
        problems.append(
            f"Toplam istenen nöbet ({total_required}) toplam kapasiteyi ({total_capacity}) aşıyor."
        )

    # 3) Kişi bazında: hedef > müsait gün sayısı
    for p in personeller:
        hedef = target_map.get(p, 0)
        musait = sum(1 for d in range(1, gun_sayisi + 1) if p in avail_by_day[d])
        if hedef > musait:
            problems.append(f"{p}: hedef {hedef}, müsait gün {musait} → hedef fazla.")

    # 4) "En az k kez birlikte" kontrolü
    for item in (want_pairs or []):
        if isinstance(item, dict):
            a, b, k = item.get("a"), item.get("b"), int(item.get("min", 0))
        elif isinstance(item, tuple) and len(item) >= 3:
            a, b, k = item[0], item[1], int(item[2])
        else:
            continue

        if a not in personeller or b not in personeller or k <= 0:
            continue

        ortak = []
        for d in range(1, gun_sayisi + 1):
            if (d not in izinler.get(a, set())) and (d not in izinler.get(b, set())):
                ortak.append(d)
        if len(ortak) < k:
            problems.append(f"{a} + {b}: en az {k} birlikte istendi ama ortak müsait gün {len(ortak)}.")

    # 5) En dar günleri işaretle
    tight_days = sorted(range(1, gun_sayisi + 1), key=lambda d: len(avail_by_day[d]))[:5]
    if tight_days and min_staff_per_day >= 2:
        problems.append(
            f"En dar günler: "
            + ", ".join([f"{d}({len(avail_by_day[d])})" for d in tight_days])
        )

    # Tatil bilgisi
    if holidays_set:
        problems.append(f"Tatil günleri: {sorted(list(holidays_set))}")

    if not problems:
        problems.append("Girdi kontrollerinde çelişki bulunamadı.")

    return problems


# ============================================================================
# STREAMLIT UI
# ============================================================================

st.set_page_config(page_title="Nöbet Planlayıcı", layout="wide")
st.title("🏥 Acil Servis Nöbet Planlayıcı")


def init_defaults():
    """Session state defaults"""
    ss = st.session_state
    ss.setdefault("yil", 2026)
    ss.setdefault("ay", 1)
    ss.setdefault("default_target", 7)
    ss.setdefault("personel_sayisi", 9)
    ss.setdefault("personel_list", ["Dr. Ahmet", "Dr. Ayşe", "Dr. Mehmet", "Dr. Fatma", "Dr. Ali", "Dr. Zeynep", "Dr. Can", "Dr. Elif", "Dr. Burak"])
    ss.setdefault("personel_targets", {})
    ss.setdefault("override_text", "")
    ss.setdefault("manual_holidays", "")  # Manuel tatil girişi
    ss.setdefault("want_pairs_list", [])
    ss.setdefault("no_pairs_list", [])
    ss.setdefault("izin_map", {})
    ss.setdefault("weekday_block_map", {})
    ss.setdefault("prefer_map", {})
    ss.setdefault("w_gap3", 300)
    ss.setdefault("soft_no_pairs_list", [])
    ss.setdefault("gap2_weight", 1000)


init_defaults()

# Sekme yapısı
tab_names = ["Kişiler", "İzin Talepleri", "Eşleşme Tercihleri", "Sonuç"]
tabs = st.tabs(tab_names)

# ============================================================================
# TAB 0: KİŞİLER
# ============================================================================
with tabs[0]:
    st.subheader("👥 Kişiler ve Hedefler")

    col1, col2, col3 = st.columns(3)
    with col1:
        st.number_input("Yıl", min_value=2020, max_value=2100, step=1, key="yil")
    with col2:
        st.number_input("Ay", min_value=1, max_value=12, step=1, key="ay")
    with col3:
        default_target = st.number_input("Varsayılan hedef nöbet", min_value=0, max_value=31, step=1, key="default_target")

    st.divider()

    personel_sayisi = st.number_input(
        "Kaç personel var?",
        min_value=1,
        max_value=50,
        value=st.session_state.get("personel_sayisi", 9),
        step=1,
        key="personel_sayisi_input"
    )

    if len(st.session_state["personel_list"]) < personel_sayisi:
        for i in range(len(st.session_state["personel_list"]), personel_sayisi):
            st.session_state["personel_list"].append(f"Personel {i+1}")
    elif len(st.session_state["personel_list"]) > personel_sayisi:
        st.session_state["personel_list"] = st.session_state["personel_list"][:personel_sayisi]

    st.session_state["personel_sayisi"] = personel_sayisi

    st.caption("Her personelin adını ve hedef nöbet sayısını girin:")

    for i in range(personel_sayisi):
        cols = st.columns([3, 1])
        with cols[0]:
            st.session_state["personel_list"][i] = st.text_input(
                f"{i+1}. Personel",
                value=st.session_state["personel_list"][i],
                key=f"personel_name_{i}"
            )
        with cols[1]:
            p_name = st.session_state["personel_list"][i]
            current_target = st.session_state.get("personel_targets", {}).get(p_name, default_target)
            new_target = st.number_input(
                "Hedef",
                min_value=0,
                max_value=31,
                value=int(current_target),
                step=1,
                key=f"target_{i}",
                help=f"Hedef nöbet sayısı (varsayılan: {default_target})"
            )
            if new_target != default_target:
                st.session_state.setdefault("personel_targets", {})[p_name] = new_target
            elif p_name in st.session_state.get("personel_targets", {}):
                if new_target == default_target:
                    st.session_state["personel_targets"].pop(p_name, None)

# ============================================================================
# TAB 1: İZİN TALEPLERİ
# ============================================================================
with tabs[1]:
    st.subheader("🏖️ İzinler")

    personeller = st.session_state.get("personel_list", [])
    if not personeller:
        st.warning("Önce Kişiler sekmesinde personel listesini gir.")
    else:
        yil = int(st.session_state["yil"])
        ay = int(st.session_state["ay"])
        gun_sayisi = calc_month_days(yil, ay)
        gun_list = list(range(1, gun_sayisi + 1))

        izin_map = st.session_state.get("izin_map", {})
        izin_map = {k: v for k, v in izin_map.items() if k in personeller}
        for p in personeller:
            izin_map.setdefault(p, [])
        st.session_state["izin_map"] = izin_map

        for p in personeller:
            with st.expander(f"📅 {p}", expanded=False):
                selected = st.multiselect(
                    label="İzinli günler",
                    options=gun_list,
                    default=sorted(list(set(st.session_state["izin_map"].get(p, [])))),
                    key=f"izin_{p}"
                )
                st.session_state["izin_map"][p] = sorted(selected)

                gun_adlari = ["Pazartesi", "Salı", "Çarşamba", "Perşembe", "Cuma", "Cumartesi", "Pazar"]
                st.session_state["weekday_block_map"].setdefault(p, [])
                blocked = st.multiselect(
                    label="Bloklu hafta günleri",
                    options=gun_adlari,
                    default=st.session_state["weekday_block_map"].get(p, []),
                    key=f"wblock_{p}"
                )
                st.session_state["weekday_block_map"][p] = blocked

                st.session_state["prefer_map"].setdefault(p, [])
                prefer_selected = st.multiselect(
                    label="Tercih edilen günler (soft)",
                    options=gun_list,
                    default=sorted(list(set(st.session_state["prefer_map"].get(p, [])))),
                    key=f"prefer_{p}"
                )
                st.session_state["prefer_map"][p] = sorted(prefer_selected)

        st.divider()
        toplam = sum(len(v) for v in st.session_state["izin_map"].values())
        st.caption(f"✓ Toplam izin günü: {toplam}")

        # Otomatik Tatil Bölümü
        st.divider()
        st.subheader("🎌 Resmi Tatiller")
        
        # Otomatik tatilleri çek
        auto_holidays = get_turkish_holidays(yil, ay)
        
        if auto_holidays:
            st.success(f"✓ Bu ay için otomatik tespit edilen tatiller:")
            for gun, isim in sorted(auto_holidays.items()):
                st.write(f"  • {gun} {datetime(yil, ay, 1).strftime('%B')[:3]} - {isim}")
        else:
            st.info("Bu ay resmi tatil bulunmuyor.")
        
        # Manuel ekleme
        st.caption("İdari izin veya ekstra tatil günü varsa ekleyin:")
        manual_input = st.text_input(
            "Ekstra tatil günleri (örn: 15, 16)",
            value=st.session_state.get("manual_holidays", ""),
            key="manual_holidays_input",
            help="Virgülle ayırarak birden fazla gün girebilirsiniz"
        )
        st.session_state["manual_holidays"] = manual_input
        
        if manual_input.strip():
            manual_days = parse_day_numbers(manual_input, gun_sayisi)
            if manual_days:
                st.caption(f"  → Eklenecek: {sorted(manual_days)}")

# ============================================================================
# TAB 2: EŞLEŞME TERCİHLERİ
# ============================================================================
with tabs[2]:
    st.subheader("👫 Eşleşme Tercihleri")

    personeller = st.session_state.get("personel_list", [])
    if not personeller:
        st.warning("Önce Kişiler sekmesinde personel listesini gir.")
    elif len(personeller) < 2:
        st.warning("Çift tanımlamak için en az 2 personel gerekli.")
    else:
        colA, colB = st.columns(2)

        with colA:
            st.markdown("### ✅ Birlikte tutsun")
            a = st.selectbox("Personel A", options=personeller, key="wp_a")
            b_options = [p for p in personeller if p != a]
            b = st.selectbox("Personel B", options=b_options, key="wp_b")
            min_k = st.number_input("Minimum birlikte gün", min_value=1, max_value=31, value=2, step=1, key="wp_min")

            if st.button("➕ Want çifti ekle", key="wp_add"):
                aa, bb = sorted([a, b])
                exists = any((item["a"] == aa and item["b"] == bb) for item in st.session_state["want_pairs_list"])
                if not exists:
                    st.session_state["want_pairs_list"].append({"a": aa, "b": bb, "min": int(min_k)})
                    st.rerun()
                else:
                    for item in st.session_state["want_pairs_list"]:
                        if item["a"] == aa and item["b"] == bb:
                            item["min"] = int(min_k)
                            st.rerun()

        with colB:
            st.markdown("### ❌ Asla birlikte tutmasın")
            na = st.selectbox("Personel A ", options=personeller, key="np_a")
            nb_options = [p for p in personeller if p != na]
            nb = st.selectbox("Personel B ", options=nb_options, key="np_b")

            if st.button("➕ No-pair ekle", key="np_add"):
                aa, bb = sorted([na, nb])
                exists = any((item["a"] == aa and item["b"] == bb) for item in st.session_state["no_pairs_list"])
                if not exists:
                    st.session_state["no_pairs_list"].append({"a": aa, "b": bb})
                    st.rerun()

        st.divider()
        st.markdown("### Mevcut tanımlar")

        colL, colR = st.columns(2)

        with colL:
            st.markdown("**Want pairs**")
            if not st.session_state["want_pairs_list"]:
                st.caption("Henüz yok.")
            else:
                for i, item in enumerate(st.session_state["want_pairs_list"]):
                    c1, c2 = st.columns([6, 2])
                    with c1:
                        st.write(f"- {item['a']} ↔ {item['b']} (min: {item['min']})")
                    with c2:
                        if st.button("Sil", key=f"wp_del_{i}"):
                            st.session_state["want_pairs_list"].pop(i)
                            st.rerun()

        with colR:
            st.markdown("**No pairs**")
            if not st.session_state["no_pairs_list"]:
                st.caption("Henüz yok.")
            else:
                for i, item in enumerate(st.session_state["no_pairs_list"]):
                    c1, c2 = st.columns([8, 2])
                    with c1:
                        st.write(f"- {item['a']} × {item['b']}")
                    with c2:
                        if st.button("Sil", key=f"np_del_{i}"):
                            st.session_state["no_pairs_list"].pop(i)
                            st.rerun()

        st.divider()
        with st.expander("⚙️ Gelişmiş Ayarlar"):
            st.info("Soft (esnek) kurallar")

            st.markdown("#### ☁️ Soft No-Pair")
            sna = st.selectbox("Personel A", options=personeller, key="snp_a")
            snb_options = [p for p in personeller if p != sna]
            snb = st.selectbox("Personel B", options=snb_options, key="snp_b")

            if st.button("➕ Esnek No-pair ekle"):
                aa, bb = sorted([sna, snb])
                exists = any((item["a"] == aa and item["b"] == bb) for item in st.session_state["soft_no_pairs_list"])
                if not exists:
                    st.session_state["soft_no_pairs_list"].append({"a": aa, "b": bb})
                    st.rerun()

            for i, item in enumerate(st.session_state["soft_no_pairs_list"]):
                sc1, sc2 = st.columns([8, 2])
                sc1.write(f"☁️ {item['a']} - {item['b']}")
                if sc2.button("Sil", key=f"snp_del_{i}"):
                    st.session_state["soft_no_pairs_list"].pop(i)
                    st.rerun()

            st.divider()
            st.markdown("#### ⚖️ Ağırlıklar")
            st.session_state.gap2_weight = st.slider(
                "Günaşırı Nöbet Cezası",
                0, 5000, st.session_state.gap2_weight
            )

            st.session_state.w_gap3 = st.slider(
                "2 Gün Boşluk Tercihi",
                0, 2000, st.session_state.w_gap3
            )

# ============================================================================
# TAB 3: ÇÖZÜM (Eski TAB 4)
# ============================================================================
with tabs[3]:
    st.subheader("✅ Çözüm")

    if st.button("🚀 Nöbeti Oluştur", type="primary", use_container_width=True):
        yil = int(st.session_state["yil"])
        ay = int(st.session_state["ay"])
        default_target = int(st.session_state["default_target"])

        personeller = st.session_state.get("personel_list", [])
        if not personeller:
            st.error("Personel listesi boş olamaz.")
            st.stop()

        gun_sayisi = calc_month_days(yil, ay)

        # Hedefler
        target = {}
        for p in personeller:
            target[p] = st.session_state.get("personel_targets", {}).get(p, default_target)

        # İzinler
        izinler = {p: set(days) for p, days in st.session_state.get("izin_map", {}).items() if days}
        prefer_map = {p: set(days) for p, days in st.session_state.get("prefer_map", {}).items() if days}

        # Hafta günü blokları
        tr_to_weekday = {
            "Pazartesi": 0, "Salı": 1, "Çarşamba": 2, "Perşembe": 3,
            "Cuma": 4, "Cumartesi": 5, "Pazar": 6
        }

        weekday_block_map = st.session_state.get("weekday_block_map", {})
        for p in personeller:
            blocked_names = weekday_block_map.get(p, [])
            if not blocked_names:
                continue
            blocked_nums = {tr_to_weekday[name] for name in blocked_names if name in tr_to_weekday}
            for d in range(1, gun_sayisi + 1):
                if datetime(yil, ay, d).weekday() in blocked_nums:
                    izinler.setdefault(p, set()).add(d)

        # Tatiller (Otomatik + Manuel)
        auto_holidays = set(get_turkish_holidays(yil, ay).keys())
        manual_holidays_text = st.session_state.get("manual_holidays", "")
        manual_holidays = parse_day_numbers(manual_holidays_text, gun_sayisi) if manual_holidays_text.strip() else set()
        holidays = auto_holidays | manual_holidays

        # Pairs
        no_pairs = [(d["a"], d["b"]) for d in st.session_state.get("no_pairs_list", [])]
        want_pairs = [(d["a"], d["b"], int(d["min"])) for d in st.session_state.get("want_pairs_list", [])]

        # Feasibility
        total_target = sum(target.values())
        if total_target < gun_sayisi:
            st.error(f"İmkânsız: Toplam hedef ({total_target}) < gün sayısı ({gun_sayisi})")
            st.stop()

        # Soft rules
        soft_no_pairs = [(d["a"], d["b"]) for d in st.session_state.get("soft_no_pairs_list", [])]
        w_gap3 = st.session_state.get("w_gap3", 300)
        gap2_weight = st.session_state.get("gap2_weight", 1000)

        st.info("Solver çalıştırılıyor...")
        try:
            schedule = solve_schedule(
                yil=yil,
                ay=ay,
                personeller=personeller,
                target=target,
                izinler=izinler,
                holidays=holidays,
                no_pairs=no_pairs,
                want_pairs=want_pairs,
                prefer_map=prefer_map,
                soft_no_pairs=soft_no_pairs,
                w_gap3=w_gap3,
                gap2_weight=gap2_weight
            )
        except Exception as e:
            st.error("❌ Çözüm bulunamadı.")
            st.caption(str(e))

            target_map = target if isinstance(target, dict) else {p: int(target) for p in personeller}

            problems = diagnose_no_solution(
                yil=yil,
                ay=ay,
                personeller=personeller,
                target_map=target_map,
                izinler=izinler,
                holidays=holidays,
                want_pairs=want_pairs,
                no_pairs=no_pairs,
                min_staff_per_day=1,
                max_staff_per_day=3,
            )

            st.warning("🔍 Olası nedenler:")
            for msg in problems:
                st.write("• " + msg)

            st.stop()

        # Sonuç tablosu
        weekdays_tr = ["Pazartesi", "Salı", "Çarşamba", "Perşembe", "Cuma", "Cumartesi", "Pazar"]

        max_kisi = max((len(v) for v in schedule.values()), default=0)
        max_kisi = max(max_kisi, 1)

        rows = []
        for d in range(1, gun_sayisi + 1):
            dt = datetime(yil, ay, d)
            wd = weekdays_tr[dt.weekday()]
            isimler = schedule.get(d, [])
            row = {
                "Gün": d,
                "Tarih": f"{d:02d}/{ay:02d}/{yil}",
                "Hafta Günü": wd,
                "Kişi Sayısı": len(isimler),
                "Tatil": "Evet" if d in set(holidays or []) else "",
            }
            for i in range(max_kisi):
                row[f"Nöbetçi {i+1}"] = isimler[i] if i < len(isimler) else ""
            rows.append(row)

        df_schedule = pd.DataFrame(rows)
        st.success("🎉 Çözüm bulundu!")
        st.subheader("📋 Oluşturulan Nöbet Listesi")

        st.dataframe(df_schedule, use_container_width=True, hide_index=True)

        csv = df_schedule.to_csv(index=False).encode('utf-8-sig')
        st.download_button(
            label="📥 CSV İndir",
            data=csv,
            file_name=f"Nobet_{yil}_{ay}.csv",
            mime="text/csv",
        )

        st.divider()

        # Nöbet dağılımı
        st.subheader("📊 Personel Nöbet Dağılımı")
        stats = []
        for p in personeller:
            count = sum(1 for d in schedule.values() if p in d)
            target_val = target.get(p, default_target)
            stats.append({
                "Personel": p,
                "Hedef": target_val,
                "Gerçekleşen": count,
                "Fark": count - target_val
            })

        st.table(pd.DataFrame(stats))

        # Excel
        fieldnames = list(rows[0].keys()) if rows else []
        xlsx_buf = io.BytesIO()
        wb = Workbook()
        ws = wb.active
        ws.title = f"Nöbet {ay:02d}-{yil}"

        header_fill = PatternFill(start_color="1F4788", end_color="1F4788", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF")
        center = Alignment(horizontal="center", vertical="center")

        for c, h in enumerate(fieldnames, start=1):
            cell = ws.cell(row=1, column=c, value=h)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = center

        fill_weekend = PatternFill(start_color="FFF4E6", end_color="FFF4E6", fill_type="solid")
        fill_holiday = PatternFill(start_color="FFE0E0", end_color="FFE0E0", fill_type="solid")

        for r_i, row in enumerate(rows, start=2):
            dt = datetime(yil, ay, row["Gün"])
            is_weekend = weekdays_tr[dt.weekday()] in ["Cuma", "Cumartesi", "Pazar"]
            is_holiday = (row["Gün"] in set(holidays or []))

            for c_i, h in enumerate(fieldnames, start=1):
                cell = ws.cell(row=r_i, column=c_i, value=row.get(h, ""))
                if c_i <= 5:
                    cell.alignment = center

                if is_holiday:
                    cell.fill = fill_holiday
                elif is_weekend:
                    cell.fill = fill_weekend

        for col in ws.columns:
            maxlen = 0
            col_letter = col[0].column_letter
            for cell in col:
                v = "" if cell.value is None else str(cell.value)
                maxlen = max(maxlen, len(v))
            ws.column_dimensions[col_letter].width = min(maxlen + 2, 30)

        wb.save(xlsx_buf)
        xlsx_buf.seek(0)

        st.download_button(
            "⬇️ Excel İndir (XLSX)",
            data=xlsx_buf.getvalue(),
            file_name=f"nobet_{ay:02d}_{yil}.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            key="dl_xlsx"
        )